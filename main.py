#!/usr/bin/env python3
"""
makeRecipt - Automatic Hebrew Receipt Generator
"""

import os
import re
import time
from datetime import datetime

import cv2
import fitz
import numpy as np
import pyautogui
import win32api, win32con, win32gui
from PIL import ImageGrab

TEMPLATE_PATH     = r"C:\Users\Administrator\Documents\קבלות_על_אימונים\example_receipt.docx"
OUTPUT_FOLDER     = r"C:\Users\Administrator\Documents\קבלות_על_אימונים"
FOXIT_PATH        = r"C:\Program Files (x86)\Foxit Software\Foxit PDF Editor\FoxitPDFEditor.exe"
EXCEL_PATH        = r"C:\Users\Administrator\Documents\קבלות_על_אימונים\training_sessions_2026.xlsx"
RECEIPT_NUM_LABEL = "קבלה מספר"
DATE_LABEL        = "תאריך"
MONTH_LABEL       = "חודש"
SUM_LABEL         = "סכום"

HEBREW_MONTHS: dict[str, str] = {
    "ינואר":"January","פברואר":"February","מרץ":"March","אפריל":"April",
    "מאי":"May","יוני":"June","יולי":"July","אוגוסט":"August",
    "ספטמבר":"September","אוקטובר":"October","נובמבר":"November","דצמבר":"December",
}
SUM_LABEL_ALTERNATIVES = [SUM_LABEL, 'סה"כ', "סה''כ", "לתשלום"]

_SIG_STRIP_Y1 = 580
_SIG_STRIP_Y2 = 650
_SIG_X_PT     = 465.0
_ABOVE_PX     = 13


def _raw_click(x, y, pause=0.3):
    win32api.SetCursorPos((x, y)); time.sleep(0.2)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0); time.sleep(0.15)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,   0, 0, 0, 0); time.sleep(pause)


def _word_replace_all(content, find_text, replace_text):
    f = content.Find; f.ClearFormatting(); f.Replacement.ClearFormatting()
    return f.Execute(find_text,False,False,False,False,False,True,1,False,replace_text,2)


def _find_label_sep_value(doc_text, label, value_pattern):
    for sep in (": ", ":", " ", ""):
        m = re.search(re.escape(label)+re.escape(sep)+f"({value_pattern})", doc_text)
        if m: return sep, m.group(1)
    return "", ""


def _update_field(content, doc_text, label, new_value, value_pattern, label_alternatives=None):
    for lbl in [label] + (label_alternatives or []):
        sep, old = _find_label_sep_value(doc_text, lbl, value_pattern)
        if old and _word_replace_all(content, lbl+sep+old, lbl+sep+new_value):
            print(f"  ✓  {lbl}: «{old}» → «{new_value}»"); return True
    print(f"  ⚠  Field not found: {label}"); return False


def _find_sigline_target(pdf_path):
    doc = fitz.open(pdf_path)
    pix = doc[0].get_pixmap(matrix=fitz.Matrix(2,2))
    rendered = np.frombuffer(pix.samples,dtype=np.uint8).reshape(pix.height,pix.width,pix.n)
    if pix.n==4: rendered=rendered[:,:,:3]
    doc.close()
    strip = rendered[_SIG_STRIP_Y1:_SIG_STRIP_Y2,:]
    scr = np.array(ImageGrab.grab(bbox=(1920,0,3840,1080),all_screens=True).convert('RGB'))
    best = None
    for scale in np.arange(0.80,1.40,0.02):
        w,h = int(strip.shape[1]*scale), int(strip.shape[0]*scale)
        if w>=scr.shape[1] or h>=scr.shape[0]: continue
        tmpl = cv2.resize(strip,(w,h))
        res  = cv2.matchTemplate(cv2.cvtColor(scr[150:],cv2.COLOR_RGB2GRAY),
                                  cv2.cvtColor(tmpl,cv2.COLOR_RGB2GRAY),cv2.TM_CCOEFF_NORMED)
        _,val,_,loc = cv2.minMaxLoc(res)
        if best is None or val>best[0]: best=(val,scale,loc[0]+1920,loc[1]+150,w,h)
    val,scale,mx,my,mw,mh = best
    print(f"  ✓  Sigline match: confidence={val:.3f}  scale={scale:.2f}")
    if val<0.50: raise RuntimeError(f"Signature line not found (confidence={val:.3f})")
    tx = mx+int(_SIG_X_PT*2*scale); ty = my+mh//2-_ABOVE_PX
    print(f"  ✓  Target: ({tx},{ty})"); return tx,ty


def _get_amount_from_excel(month_en: str) -> str:
    """
    Open training_sessions_2026.xlsx, go to the sheet named *month_en*,
    find the cell containing 'הכנסה צפויה:' and return the value from the
    cell immediately to its left (one column lower index).
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run:  pip install openpyxl")

    if not os.path.exists(EXCEL_PATH):
        raise RuntimeError(f"Excel file not found:\n  {EXCEL_PATH}")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    try:
        if month_en not in wb.sheetnames:
            raise RuntimeError(
                f"Sheet '{month_en}' not found in workbook.\n"
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb[month_en]
        target = "הכנסה צפויה:"
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.strip() == target:
                    if cell.column <= 1:
                        raise RuntimeError(
                            f"Found '{target}' at {cell.coordinate} but there is no cell to its left."
                        )
                    left_val = ws.cell(row=cell.row, column=cell.column - 1).value
                    if left_val is None or str(left_val).strip() == "":
                        raise RuntimeError(
                            f"Cell to the left of {cell.coordinate} ('{target}') is empty."
                        )
                    # Strip commas / currency symbols if present, keep as string
                    return str(left_val).replace(",", "").strip()
        raise RuntimeError(f"'{target}' not found in sheet '{month_en}'.")
    finally:
        wb.close()


def apply_signature_in_foxit(pdf_path):
    try: import pywinauto
    except ImportError: print("  ⚠  pywinauto not installed."); return False

    # ── 0. Wait for Foxit to load the PDF ────────────────────────
    pdf_name = os.path.basename(pdf_path)
    app = None
    for _ in range(30):
        try:
            a = pywinauto.Application(backend='uia').connect(path=FOXIT_PATH, timeout=3)
            for w in a.windows():
                if pdf_name in w.window_text():
                    app = a
                    break
            if app:
                break
        except Exception:
            pass
        time.sleep(1)
    if not app:
        print("  ⚠  Foxit did not open the PDF within 30 s")
        return False
    print("PDF loaded")

    # ── 1. Move to right monitor + maximize + verify ─────────────
    win  = app.top_window()
    hwnd = win.handle
    # Restore out of any maximized state first
    win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
    time.sleep(0.8)
    # Move window deep inside the right monitor (x 1920-3840) then maximize
    win32gui.MoveWindow(hwnd, 2500, 200, 600, 400, True)
    time.sleep(0.5)
    win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
    win32gui.BringWindowToTop(hwnd)
    time.sleep(3.0)

    win = app.top_window()
    r   = win.rectangle()
    wl, wt, wr, wb = r.left, r.top, r.right, r.bottom
    print(f"Window: ({wl},{wt}) -> ({wr},{wb})  size={wr-wl}x{wb-wt}")
    if (wr - wl) < 1600 or (wb - wt) < 800:
        print("Window is not full screen — maximize failed.")
        return False
    print("Full screen confirmed")

    # ── 2. Click Home tab ─────────────────────────────────────────
    home = [t for t in win.descendants(control_type='TabItem') if t.window_text() == 'Home']
    if not home:
        print("'Home' tab not found via UIA")
        return False
    home[0].click_input()
    time.sleep(0.8)
    print("Home tab clicked")

    # ── 2b. Set zoom to 100% ──────────────────────────────────────
    win = app.top_window()
    zoom_el = [el for el in win.descendants(control_type='Edit')
               if el.rectangle().left > 3600 and el.rectangle().top > 980]
    if zoom_el:
        zoom_el[-1].double_click_input()
        time.sleep(0.4)
        pyautogui.typewrite('100', interval=0.05)
        pyautogui.press('enter')
        time.sleep(0.8)
        print("Zoom set to 100%")
    else:
        print("  (zoom edit not found, continuing)")

    # ── 3. Page Up x8 to reach document top ──────────────────────
    win32gui.BringWindowToTop(hwnd)
    time.sleep(0.3)
    for _ in range(8):
        pyautogui.press('pageup')
        time.sleep(0.15)
    time.sleep(0.5)
    print("Scrolled to document top")

    # ── 4. Locate the dotted signature line via template matching ─
    doc = fitz.open(pdf_path)
    pix = doc[0].get_pixmap(matrix=fitz.Matrix(2, 2))
    rendered = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
    if pix.n == 4:
        rendered = rendered[:, :, :3]
    doc.close()
    strip = rendered[_SIG_STRIP_Y1:_SIG_STRIP_Y2, :]

    scr = np.array(ImageGrab.grab(bbox=(1920, 0, 3840, 1080), all_screens=True).convert('RGB'))
    best = None
    for scale in np.arange(0.80, 1.40, 0.02):
        w = int(strip.shape[1] * scale)
        h = int(strip.shape[0] * scale)
        if w >= scr.shape[1] or h >= scr.shape[0]:
            continue
        tmpl = cv2.resize(strip, (w, h))
        res  = cv2.matchTemplate(
            cv2.cvtColor(scr[150:], cv2.COLOR_RGB2GRAY),
            cv2.cvtColor(tmpl,      cv2.COLOR_RGB2GRAY),
            cv2.TM_CCOEFF_NORMED)
        _, val, _, loc = cv2.minMaxLoc(res)
        if best is None or val > best[0]:
            best = (val, scale, loc[0] + 1920, loc[1] + 150, w, h)

    val, scale, mx, my, mw, mh = best
    print(f"Sigline match: confidence={val:.3f}  scale={scale:.2f}")
    print(f"Match top-left on screen: ({mx},{my})  template size: {mw}x{mh}")
    if val < 0.50:
        print(f"Signature line not found on screen (confidence={val:.3f} < 0.50).")
        return False

    tx = mx + int(_SIG_X_PT * 2 * scale)
    ty = my + mh // 2 - _ABOVE_PX
    print(f"Target: ({tx},{ty})")

    # ── 5. Click Fill & Sign (Button in fresh window, Tab if already used) ──
    win = app.top_window()
    fs = [t for t in win.descendants(control_type='TabItem') if t.window_text() == 'Fill & Sign']
    if fs:
        fs[0].click_input()
        print("Fill & Sign tab clicked")
    else:
        btn = [b for b in win.descendants(control_type='Button')
               if 'Fill' in b.window_text() and 'Sign' in b.window_text()]
        if not btn:
            print("'Fill & Sign' button/tab not found via UIA")
            return False
        btn[0].click_input()
        print(f"Fill & Sign button clicked: [{btn[0].window_text()}]")
    time.sleep(2.0)

    # ── 6. Find Signature List button (retry up to 8 s) ──────────
    sig_btn = None
    for attempt in range(8):
        win = app.top_window()
        sig_btn = [b for b in win.descendants(control_type='Button')
                   if b.window_text() == 'Signature List']
        if sig_btn:
            break
        print(f"  Signature List not ready... ({attempt+1}/8)")
        time.sleep(1.0)
    if not sig_btn:
        print("'Signature List' button not found via UIA after 8 s")
        return False

    br    = sig_btn[0].rectangle()
    sig_x = br.left + 80
    sig_y = (br.top + br.bottom) // 2
    print(f"Signature List rect: ({br.left},{br.top},{br.right},{br.bottom})  clicking at ({sig_x},{sig_y})")

    # ── 7. Pick up signature + place it ──────────────────────────
    _raw_click(sig_x, sig_y, pause=1.5)
    _raw_click(tx,    ty,    pause=1.5)

    # ── 8. Move mouse away (no click, no Escape) ──────────────────
    win32api.SetCursorPos((wl + 150, wt + 300))
    time.sleep(0.5)

    # ── 9. Apply All Signatures ───────────────────────────────────
    win = app.top_window()
    apply_btn = [b for b in win.descendants(control_type='Button')
                 if b.window_text() == 'Apply All Signatures']
    if not apply_btn:
        print("'Apply All Signatures' button not found via UIA")
        return False
    apply_btn[0].click_input()
    time.sleep(2.0)
    print("Apply All Signatures clicked")

    # ── 10. Save ──────────────────────────────────────────────────
    win32gui.BringWindowToTop(hwnd)
    time.sleep(0.3)
    pyautogui.hotkey('ctrl', 's')
    time.sleep(2.0)
    pyautogui.press('enter')
    time.sleep(2.0)
    print("Saved")

    # ── 11. Close Foxit ───────────────────────────────────────────
    pyautogui.hotkey('alt', 'f4')
    time.sleep(1.5)
    pyautogui.press('enter')
    time.sleep(1.0)
    print("Foxit closed — all done!")
    return True


def main():
    if not os.path.exists(TEMPLATE_PATH):
        print(f"Error: template not found —\n  {TEMPLATE_PATH}"); return

    print("="*55); print("      Automatic Receipt Generator"); print("="*55)
    while True:
        month_he = input("\nEnter month (Hebrew): ").strip()
        if month_he in HEBREW_MONTHS:
            break
        print("Unknown month. Options:\n  " + "  |  ".join(HEBREW_MONTHS))

    month_en = HEBREW_MONTHS[month_he]
    try:
        suggested = _get_amount_from_excel(month_en)
        print(f"Amount loaded from Excel ({month_en} sheet): {suggested}")
    except Exception as exc:
        print(f"  (Could not read amount from Excel: {exc})")
        suggested = ""

    user_input = input(f"Enter amount [{suggested}]: ").strip()
    amount = user_input if user_input else suggested
    if not amount:
        print("Error: no amount provided."); return
    print(f"Amount: {amount}")

    today    = datetime.now()
    date_str = today.strftime("%d/%m/%Y")
    year     = today.year

    try: import win32com.client
    except ImportError: print("Error: pywin32 not installed."); return

    word=win32com.client.Dispatch("Word.Application"); word.Visible=False
    abs_template=os.path.abspath(TEMPLATE_PATH)
    try:
        doc=word.Documents.Open(abs_template)
        doc_text=doc.Content.Text; content=doc.Content
        print("\nUpdating document fields:")

        sep_r,old_num_str=_find_label_sep_value(doc_text,RECEIPT_NUM_LABEL,r"\d+")
        if not old_num_str: print(f"  Error: '{RECEIPT_NUM_LABEL}' not found."); doc.Close(False); return
        new_num=int(old_num_str)+1
        _word_replace_all(content,RECEIPT_NUM_LABEL+sep_r+old_num_str,RECEIPT_NUM_LABEL+sep_r+str(new_num))
        print(f"  ✓  {RECEIPT_NUM_LABEL}: {old_num_str} → {new_num}")

        _update_field(content,doc_text,DATE_LABEL,date_str,r"\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4}")
        _update_field(content,doc_text,MONTH_LABEL,month_he,"|".join(re.escape(m) for m in HEBREW_MONTHS))
        _update_field(content,doc_text,SUM_LABEL,amount,r"[\d,\.]+",label_alternatives=SUM_LABEL_ALTERNATIVES[1:])

        doc.Save(); print(f"\n✓  Document saved:  {abs_template}")

        os.makedirs(OUTPUT_FOLDER,exist_ok=True)
        pdf_filename=f"recipt_{new_num}_{month_en}_{year}.pdf"
        pdf_path=os.path.abspath(os.path.join(OUTPUT_FOLDER,pdf_filename))
        doc.ExportAsFixedFormat(pdf_path,17)
        print(f"✓  PDF saved:      {pdf_path}"); doc.Close(False)

        if os.path.exists(FOXIT_PATH):
            from place_sig_once import apply_signature
            print("\nApplying signature...")
            apply_signature(pdf_path, log_fn=print)
        else:
            print(f"  ⚠  Foxit not found at: {FOXIT_PATH}")

    except Exception as exc:
        print(f"\nError: {exc}"); raise
    finally:
        try: word.Quit()
        except: pass

    print("\n✓  All done!")


if __name__ == "__main__":
    main()
